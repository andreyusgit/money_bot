import asyncio
import logging
from contextlib import suppress
import os.path
import openpyxl
import markdown

from aiogram.utils.exceptions import MessageCantBeDeleted, MessageToDeleteNotFound
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

from aiogram import Bot, types
from aiogram.utils import executor
from aiogram.dispatcher import Dispatcher
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.contrib.middlewares.logging import LoggingMiddleware
from aiogram.types import ReplyKeyboardRemove, \
    ReplyKeyboardMarkup, KeyboardButton, \
    InlineKeyboardMarkup, InlineKeyboardButton

from config import TOKEN
from utils import TestStates
from messages import MESSAGES

logging.basicConfig(format=u'%(filename)+13s [ LINE:%(lineno)-4s] %(levelname)-8s [%(asctime)s] %(message)s',
                    level=logging.DEBUG)

bot = Bot(token=TOKEN)
dp = Dispatcher(bot, storage=MemoryStorage())
dp.middleware.setup(LoggingMiddleware())
all_users = openpyxl.Workbook()
card_sheet = all_users.create_sheet("Sheet_A")
card_sheet.title = "cards"
user_deb = {}
user_data = {}
groups = {}
members = {}


async def delete_message(message: types.Message, sleep_time: int = 0):
    await asyncio.sleep(sleep_time)
    with suppress(MessageCantBeDeleted, MessageToDeleteNotFound):
        await message.delete()


@dp.callback_query_handler(text='Yes')
async def process_callback_button1(callback_query: types.CallbackQuery):
    name = callback_query.from_user.username
    check = user_data[name].pop()
    val = float(user_data[name].pop())
    name_table = user_data[name].pop()
    username = user_data[name].pop()
    if check == 0:
        val = -val
        temp = username
        username = name
        name = temp
    wb = openpyxl.load_workbook(f'{name_table}.xlsx')
    main_sheet = wb['main']
    rows = main_sheet.max_row + 1
    coll = 1
    for i in range(2, rows):
        if main_sheet[f'{get_column_letter(i)}1'].value is None:
            break
        if main_sheet[f'{get_column_letter(i)}1'].value == username:
            coll = i
            break
    letter_col = get_column_letter(coll)
    for i in range(2, rows):
        if main_sheet[f'A{i}'].value is None:
            break
        if main_sheet[f'A{i}'].value == name:
            main_sheet[f'{letter_col}{i}'] = main_sheet[f'{letter_col}{i}'].value + val
            if main_sheet[f'{get_column_letter(i)}{coll}'].value > 0:
                if main_sheet[f'{letter_col}{i}'].value > \
                        main_sheet[f'{get_column_letter(i)}{coll}'].value:
                    main_sheet[f'{letter_col}{i}'] = main_sheet[f'{letter_col}{i}'].value - \
                                                                  main_sheet[f'{get_column_letter(i)}{coll}'].value
                    main_sheet[f'{get_column_letter(i)}{coll}'] = 0
                else:
                    main_sheet[f'{get_column_letter(i)}{coll}'] = main_sheet[f'{get_column_letter(i)}{coll}'].value - \
                                                                  main_sheet[f'{letter_col}{i}'].value
                    main_sheet[f'{letter_col}{i}'] = 0
            else:
                if main_sheet[f'{letter_col}{i}'].value < 0:
                    main_sheet[f'{get_column_letter(i)}{coll}'] = -main_sheet[f'{letter_col}{i}'].value
                    main_sheet[f'{letter_col}{i}'] = 0

            wb.save(f"{name_table}.xlsx")
            break
    if check == 1:
        await bot.send_message(callback_query.from_user.id, 'Долг успешно записан')
    elif check == 0:
        await bot.send_message(callback_query.from_user.id, 'Возврат успешно произведен')


@dp.callback_query_handler(text='No')
async def process_callback_button1(callback_query: types.CallbackQuery):
    name = callback_query.from_user.username
    check = user_data[name].pop()
    username = user_data[name].pop(0)
    id_for_mess = 0
    wb = openpyxl.load_workbook('users.xlsx')
    active_sheet = wb['Sheet']
    rows = active_sheet.max_row + 1
    for i in range(1, rows):
        if name == active_sheet[f'A{i}'].value:
            id_for_mess = int(active_sheet[f'F{i}'].value)
            break
    if check == 1:
        await bot.send_message(id_for_mess,
                               f'Долг не был подтвержден со стороны @{username}, рекомендуем связаться с этим '
                               f'пользователем в личных сообщениях и выяснить причину, *а после снова завести долг*',
                               parse_mode="Markdown")
    elif check == 0:
        await bot.send_message(id_for_mess,
                               f'Возврат не был подтвержден со стороны @{username}, рекомендуем связаться с этим '
                               f'пользователем в личных сообщениях и выяснить причину, *а после снова завести возврат*',
                               parse_mode="Markdown")


async def check_user(username, name, name_table, value, check):
    inline_btn_1 = InlineKeyboardButton('Да', callback_data='Yes')
    inline_btn_2 = InlineKeyboardButton('Нет', callback_data='No')
    inline_kb1 = InlineKeyboardMarkup().add(inline_btn_1, inline_btn_2)
    title = ''
    id_for_mess = 0
    wb = openpyxl.load_workbook('users.xlsx')
    active_sheet = wb['Sheet']
    rows = active_sheet.max_row + 1
    for i in range(1, rows):
        if name_table == active_sheet[f'B{i}'].value:
            title = active_sheet[f'C{i}'].value
            break
    for i in range(1, rows):
        if username == active_sheet[f'A{i}'].value:
            id_for_mess = int(active_sheet[f'F{i}'].value)
            break
    if check == 1:
        await bot.send_message(id_for_mess,
                               f'Пользователь @{name} из группы {title} *завел на тебя долг* в размере {value} '
                               f'рублей, если все верно - нажми на кнопку "Да", если же что-то не так, жми на '
                               f'кнопку "Нет"', reply_markup=inline_kb1, parse_mode="Markdown")
    elif check == 0:
        await bot.send_message(id_for_mess,
                               f'Пользователь @{name} из группы {title} *вернул тебе долг* в размере {value} '
                               f'рублей, если все верно - нажми на кнопку "Да", если же что-то не так, жми на '
                               f'кнопку "Нет"', reply_markup=inline_kb1, parse_mode="Markdown")

    user_data[username] = [name, name_table, value, check]


@dp.message_handler(state='*', content_types=["new_chat_members"])
async def new_member(message: types.Message):
    await message.delete()
    await message.answer(MESSAGES['hello'])


@dp.message_handler(state='*', commands=['add_me'])
async def process_add_user_command(message: types.Message):
    if str(message.chat.type) == 'group':
        username = message.from_user.username
        name_table = str(abs(message.chat.id))
        if "/add_me" in message.text:
            await message.delete()
        wb = openpyxl.load_workbook(f'{name_table}.xlsx')
        first_sheet = wb['main']
        rows = first_sheet.max_row + 1
        if rows == 2:
            rows += 1
        for i in range(2, rows + 1):
            if first_sheet[f'A{i}'].value == username:
                msg = await bot.send_message(message.chat.id, f'@{username} ты уже есть в базе')
                asyncio.create_task(delete_message(msg, 5))
                break
            elif first_sheet[f'A{i}'].value is None:
                coll = get_column_letter(i)
                first_sheet[f'A{i}'] = username
                first_sheet[f'{coll}1'] = username
                for j in range(2, i + 1):
                    first_sheet[f'{get_column_letter(j)}{i}'] = 0
                    first_sheet[f'{get_column_letter(i)}{j}'] = 0
                wb.save(f"{abs(int(name_table))}.xlsx")
                users = openpyxl.load_workbook('users.xlsx')
                sheet = users['Sheet']
                row_count = sheet.max_row + 1
                sheet[f'A{row_count}'] = username
                sheet[f'B{row_count}'] = name_table
                sheet[f'C{row_count}'] = message.chat.title
                sheet[f'D{row_count}'] = message.from_user.first_name
                sheet[f'E{row_count}'] = message.from_user.last_name
                sheet[f'F{row_count}'] = message.from_user.id
                users.save('users.xlsx')
                break


@dp.message_handler(commands=['start_me'])
async def process_start_command(message: types.Message):
    if str(message.chat.type) == 'group':
        name_table = str(abs(message.chat.id))
        abspath = os.path.abspath(f"{abs(int(name_table))}.xlsx")
        abspath_2 = os.path.abspath("users.xlsx")
        if not os.path.exists(abspath_2):
            try:
                worksheet = all_users['Sheet']
                worksheet['A1'] = '---'
                all_users.save('users.xlsx')
            except Exception as ex:
                await bot.send_message(message.chat.id, str(ex))
        if not os.path.exists(abspath):
            try:
                wb = Workbook()
                ws1 = wb.create_sheet("main")
                ws1.title = "main"
                wb.remove_sheet(wb["Sheet"])
                wb.save(f"{abs(int(name_table))}.xlsx")
            except BaseException as err:
                await bot.send_message(message.chat.id, f"Unexpected {err=}, {type(err)=}")
        await bot.send_message(message.chat.id, MESSAGES['hello'])


@dp.message_handler(commands=['my_debts', 'my_debtors'])
async def process_add_user_command(message: types.Message):
    if str(message.chat.type) == 'private':
        abspath = os.path.abspath("")
        all_tables = []
        mes = ''
        username = message.from_user.username
        for root, dirs, files in os.walk(abspath):
            for file in files:
                if file.endswith(".xlsx"):
                    if file.strip(".xlsx") != 'user':
                        all_tables.append(file.strip(".xlsx"))
        length = len(all_tables)
        if message.get_command(message) == 'my_debtors':
            mes = 'МОИ ДОЛЖНИКИ:\n\n'
        elif message.get_command(message) == 'my_debts':
            mes = 'МОИ ДОЛГИ:\n\n'
        for index in range(length):
            name_table = all_tables.pop()
            wb = openpyxl.load_workbook('users.xlsx')
            active_sheet = wb['Sheet']
            r = active_sheet.max_row + 1
            title = ''
            for i in range(1, r):
                if name_table == active_sheet[f'B{i}'].value:
                    title = active_sheet[f'C{i}'].value
                    break
            wb = openpyxl.load_workbook(f'{name_table}.xlsx')
            current_sheet = wb['main']
            if message.get_command(message) == 'my_debtors':
                rows = current_sheet.max_row + 1
                coll = -1
                for i in range(2, rows):
                    if current_sheet[f'{get_column_letter(i)}1'].value is None:
                        break
                    if current_sheet[f'{get_column_letter(i)}1'].value == username:
                        coll = get_column_letter(i)
                        break
                if coll != -1:
                    for i in range(2, rows):
                        if current_sheet[f'{coll}{i}'].value is None:
                            break
                        if current_sheet[f'{coll}{i}'].value != 0:
                            val = current_sheet[f'{coll}{i}'].value
                            mes = mes + '@' + current_sheet[f'A{i}'].value + ': ' + f"{val:.{2}f}" + \
                                  f' рублей (из чата {title})' + '\n'
            elif message.get_command(message) == 'my_debts':
                row = -1
                for i in range(2, 1000):
                    if current_sheet[f'A{i}'].value is None:
                        break
                    if current_sheet[f'A{i}'].value == username:
                        row = i
                        break
                if row != -1:
                    for i in range(2, 1000):
                        if current_sheet[f'{get_column_letter(i)}{row}'].value is None:
                            break
                        if current_sheet[f'{get_column_letter(i)}{row}'].value != 0:
                            val = current_sheet[f'{get_column_letter(i)}{row}'].value
                            mes = mes + '@' + current_sheet[f'{get_column_letter(i)}1'].value + ': ' + f"{val:.{2}f}" + \
                                  f' рублей (из чата {title})' + '\n'
        await bot.send_message(message.from_user.id, mes)


@dp.message_handler(state='*', commands=['add_debts', 'delete_debts'])
async def process_add_debts_groups_command(message: types.Message):
    if str(message.chat.type) == 'private':
        groups_keyboard = types.ReplyKeyboardMarkup(resize_keyboard=False, one_time_keyboard=True)
        username = message.from_user.username
        wb2 = openpyxl.load_workbook('users.xlsx')
        active_sheet = wb2['Sheet']
        check = 1
        if message.get_command(message) == 'delete_debts':
            check = 0
        user_deb[username] = [], [], check
        groups[username] = []
        for i in range(1, active_sheet.max_row + 1):
            if username == active_sheet[f'A{i}'].value:
                value = active_sheet[f'C{i}'].value
                user_deb[username][0].append(value)
                groups[username].append(value)
        for i in range(len(user_deb[username][0])):
            groups_keyboard.add(f'{user_deb[username][0][i]}')
        if check == 1:
            await bot.send_message(message.chat.id, 'Выбери из списка ниже в каком чате находится твой должник: ',
                                   reply_markup=groups_keyboard)
        else:
            await bot.send_message(message.chat.id, 'В каком чате находится человек, которому ты хочешь вернуть долг: ',
                                   reply_markup=groups_keyboard)
        state = dp.current_state(user=message.from_user.id)
        await state.set_state(TestStates.all()[1])


@dp.message_handler(state=TestStates.TEST_STATE_1)
async def process_add_debts_user_command(message: types.Message):
    username = message.from_user.username
    state = dp.current_state(user=message.from_user.id)
    try:
        index_value = groups[username].index(message.text)
        await state.set_state(TestStates.all()[2])
    except ValueError:
        markup = types.ReplyKeyboardRemove()
        await bot.send_message(message.chat.id, 'Такой группы нет', reply_markup=markup)
        await state.reset_state(with_data=False)
        return
    members[username] = []
    wb2 = openpyxl.load_workbook('users.xlsx')
    active_sheet = wb2['Sheet']
    name_table = ''
    member_keyboard = types.ReplyKeyboardMarkup(resize_keyboard=False, one_time_keyboard=True)
    for i in range(1, active_sheet.max_row + 1):
        if message.text == active_sheet[f'C{i}'].value:
            name_table = str(active_sheet[f'B{i}'].value)
    user_deb[message.from_user.username][1].append(name_table)
    wb = openpyxl.load_workbook(f'{name_table}.xlsx')
    main_sheet = wb['main']
    for i in range(2, main_sheet.max_row + 1):
        value = main_sheet[f'A{i}'].value
        if value != username:
            member_keyboard.add(value)
            members[username].append(value)
    check = user_deb[message.from_user.username][2]
    if check == 1:
        await bot.send_message(message.chat.id, 'Выбери из списка ниже пользователя, который тебе задолжал: ',
                               reply_markup=member_keyboard)
    elif check == 0:
        await bot.send_message(message.chat.id, 'Выбери из списка ниже пользователя, которому ты хочешь вернуть долг: ',
                               reply_markup=member_keyboard)


@dp.message_handler(state=TestStates.TEST_STATE_2)
async def process_add_debts_money_command(message: types.Message):
    username = message.from_user.username
    state = dp.current_state(user=message.from_user.id)
    try:
        index_value = members[username].index(message.text)
        await state.set_state(TestStates.all()[3])
    except ValueError:
        markup = types.ReplyKeyboardRemove()
        await bot.send_message(message.chat.id, 'Такого пользователя нет', reply_markup=markup)
        await state.reset_state(with_data=False)
        return
    user_deb[username][1].append(message.text)
    markup = types.ReplyKeyboardRemove()
    check = user_deb[username][2]
    if check == 1:
        await bot.send_message(message.chat.id, 'введи сумму долга:', reply_markup=markup)
    elif check == 0:
        await bot.send_message(message.chat.id, 'введи сумму возврата:', reply_markup=markup)


@dp.message_handler(state=TestStates.TEST_STATE_3)
async def process_add_debts_total_command(message: types.Message):
    try:
        value = [float(i) for i in message.text.replace(',', '.').split()].pop()
        username = user_deb[message.from_user.username][1].pop()
        name_table = user_deb[message.from_user.username][1].pop()
        check = user_deb[message.from_user.username][2]
        if check == 1:
            await bot.send_message(message.chat.id,
                                   f"Долг в {value} рублей будет записан на пользователя @{username} "
                                   f"после подтверждения со стороны пользователя")
        elif check == 0:
            await bot.send_message(message.chat.id,
                                   f"Долг в размере {value} рублей будет погашен после подтверждения со стороны"
                                   f" пользователя {username}")

        await check_user(username, message.from_user.username, name_table, value, check)

    except ValueError:
        await bot.send_message(message.chat.id, "Вы не ввели число")
    except BaseException as err:
        await bot.send_message(message.chat.id, f"Unexpected {err=}, {type(err)=}")
    state = dp.current_state(user=message.from_user.id)
    await state.reset_state(with_data=False)


@dp.message_handler(state='*', commands=['start'])
async def process_start_command(message: types.Message):
    await message.answer(MESSAGES['start'])


@dp.message_handler(state='*', commands=['help'])
async def process_help_command(message: types.Message):
    await message.answer(MESSAGES['help'])


@dp.message_handler()
async def some_test_state_case_met(message: types.Message):
    await message.answer(MESSAGES['no_command'])


@dp.message_handler(state='*', commands=['stepa_hvatit'])
async def echo_message(msg: types.Message):
    await bot.send_message(msg.chat.id, "остановись")
    await bot.send_message(765839138, "хватит....")


async def shutdown(dispatcher: Dispatcher):
    await dispatcher.storage.close()
    await dispatcher.storage.wait_closed()


if __name__ == '__main__':
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    button_1 = types.KeyboardButton(text="/start")
    keyboard.add(button_1)
    executor.start_polling(dp, on_shutdown=shutdown)
